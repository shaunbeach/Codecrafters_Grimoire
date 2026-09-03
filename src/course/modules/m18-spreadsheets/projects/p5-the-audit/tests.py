from openpyxl import load_workbook

SALES = "sales.xlsx"


def test_finds_both_problem_rows():
    """Two rows need attention, and they are named"""
    reset_sales()
    run_audit = require("run_audit")
    set_input(["skip", "skip"])
    clear_output()
    result = run_audit(SALES)
    printed = get_output()

    assert result["checked"] == 2, (
        f"checked came out as {result['checked']}. Two rows have no units: East/February "
        "has 0 and West/February is empty."
    )
    assert "Row 7: East February has no units." in printed, f"Your output was:\n{printed}"
    assert "Row 9: West February has no units." in printed, (
        f"Your output was:\n{printed}\nRow numbers are spreadsheet rows, counting the header."
    )


def test_a_number_fixes_the_row():
    """A whole number is written into the cell"""
    reset_sales()
    run_audit = require("run_audit")
    set_input(["45", "60"])
    result = run_audit(SALES)

    assert result == {"checked": 2, "fixed": 2, "skipped": 0}, f"Got {result}"
    sheet = load_workbook(SALES).active
    assert sheet.cell(row=7, column=3).value == 45, (
        f"Row 7 holds {sheet.cell(row=7, column=3).value!r} — the answer should be written "
        "into the Units cell as a number."
    )
    assert sheet.cell(row=9, column=3).value == 60


def test_skip_leaves_the_cell_alone():
    """'skip' changes nothing"""
    reset_sales()
    run_audit = require("run_audit")
    set_input(["skip", "skip"])
    result = run_audit(SALES)

    assert result == {"checked": 2, "fixed": 0, "skipped": 2}, f"Got {result}"
    sheet = load_workbook(SALES).active
    assert sheet.cell(row=9, column=3).value is None, (
        "A skipped row should be exactly as it was."
    )


def test_nonsense_does_not_crash():
    """A typo is a skip, not a crash"""
    reset_sales()
    run_audit = require("run_audit")
    set_input(["banana", "12"])
    try:
        result = run_audit(SALES)
    except ValueError:
        raise AssertionError(
            "run_audit died on 'banana'. int() raises ValueError — catch it and treat "
            "the answer as a skip."
        )
    assert result == {"checked": 2, "fixed": 1, "skipped": 1}, f"Got {result}"


def test_blank_answer_is_a_skip():
    """Pressing Enter skips the row"""
    reset_sales()
    run_audit = require("run_audit")
    set_input(["", ""])
    assert require("run_audit")(SALES)["skipped"] == 2, "An empty answer is not a number."


def test_negative_is_a_skip():
    """Negative units are refused"""
    reset_sales()
    run_audit = require("run_audit")
    set_input(["-5", "skip"])
    result = run_audit(SALES)
    assert result["fixed"] == 0, f"Got {result} — you cannot sell minus five things."
    sheet = load_workbook(SALES).active
    assert sheet.cell(row=7, column=3).value in (0, None)


def test_prompt_is_used():
    """The question is asked with input()"""
    reset_sales()
    run_audit = require("run_audit")
    set_input(["skip", "skip"])
    clear_output()
    run_audit(SALES)
    assert "units>" in get_output(), (
        f"The prompt never reached the screen. Pass it to input(): input('units> ')."
    )


def test_summary_and_save():
    """It reports what it did, and the change survives"""
    reset_sales()
    run_audit = require("run_audit")
    set_input(["45", "skip"])
    clear_output()
    run_audit(SALES)

    assert "Audit complete: 1 fixed, 1 skipped." in get_output(), (
        f"Your output was:\n{get_output()}"
    )
    fresh = load_workbook(SALES).active
    assert fresh.cell(row=7, column=3).value == 45, (
        "The fix is not on disk. book.save(path) after the loop."
    )
