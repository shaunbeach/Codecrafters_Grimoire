from openpyxl import load_workbook


def run_audit(path):
    book = load_workbook(path)
    sheet = book.active

    checked = 0
    fixed = 0
    skipped = 0

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=3).value:
            continue

        checked += 1
        region = sheet.cell(row=row, column=1).value
        month = sheet.cell(row=row, column=2).value
        print(f"Row {row}: {region} {month} has no units.")

        answer = input("units> ")
        try:
            units = int(answer)
        except ValueError:
            skipped += 1
            continue

        if units < 0:
            skipped += 1
            continue

        sheet.cell(row=row, column=3, value=units)
        fixed += 1

    book.save(path)
    print(f"Audit complete: {fixed} fixed, {skipped} skipped.")
    return {"checked": checked, "fixed": fixed, "skipped": skipped}
