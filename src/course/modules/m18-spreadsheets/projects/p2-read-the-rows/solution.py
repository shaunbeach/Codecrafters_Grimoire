from openpyxl import load_workbook


def load_sales(path):
    book = load_workbook(path)
    sheet = book.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = rows[0]
    records = []
    for row in rows[1:]:
        record = dict(zip(headers, row))
        record["Units"] = record["Units"] or 0
        records.append(record)
    return records
