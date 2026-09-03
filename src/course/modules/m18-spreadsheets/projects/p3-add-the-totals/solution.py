from openpyxl import load_workbook


def add_totals(path):
    book = load_workbook(path)
    sheet = book.active

    total_column = sheet.max_column + 1
    sheet.cell(row=1, column=total_column, value="Total")

    written = 0
    for row in range(2, sheet.max_row + 1):
        units = sheet.cell(row=row, column=3).value or 0
        price = sheet.cell(row=row, column=4).value or 0
        sheet.cell(row=row, column=total_column, value=units * price)
        written += 1

    book.save(path)
    return written
