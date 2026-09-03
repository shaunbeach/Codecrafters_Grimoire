from openpyxl import load_workbook

SALES = "sales.xlsx"


def _sheet():
    return load_workbook(SALES).active


def test_returns_row_count():
    """add_totals reports how many rows it filled"""
    reset_sales()
    add_totals = require("add_totals")
    result = add_totals(SALES)
    assert result == 8, (
        f"There are 8 data rows; you returned {result!r}. Count the rows you wrote, "
        "not the rows in the sheet."
    )
    assert isinstance(result, int), f"Expected an int, got {type(result).__name__}."


def test_header_written():
    """The new column is labelled Total"""
    reset_sales()
    require("add_totals")(SALES)
    sheet = _sheet()
    assert sheet.cell(row=1, column=5).value == "Total", (
        f"E1 holds {sheet.cell(row=1, column=5).value!r}. The header goes in the first "
        "empty column, which is one past max_column."
    )


def test_totals_are_correct():
    """Each row is Units multiplied by Unit Price"""
    reset_sales()
    require("add_totals")(SALES)
    sheet = _sheet()
    expected = [540.0, 427.5, 360.0, 700.0, 300.0, 0.0, 787.5, 0.0]
    for index, want in enumerate(expected, start=2):
        got = sheet.cell(row=index, column=5).value
        assert got == want, f"E{index} should be {want}, got {got!r}"


def test_blank_units_total_zero():
    """The row with no Units still gets a total"""
    reset_sales()
    require("add_totals")(SALES)
    sheet = _sheet()
    assert sheet.cell(row=9, column=5).value == 0.0, (
        f"E9 is {sheet.cell(row=9, column=5).value!r}. That row has an empty Units cell — "
        "None cannot be multiplied, so treat it as 0."
    )


def test_header_not_overwritten():
    """The first total does not land on the header"""
    reset_sales()
    require("add_totals")(SALES)
    sheet = _sheet()
    value = sheet.cell(row=1, column=5).value
    assert value == "Total", (
        f"E1 holds {value!r}. Data rows start at row 2 — range(2, max_row + 1)."
    )


def test_original_columns_untouched():
    """Nothing else in the sheet changed"""
    reset_sales()
    require("add_totals")(SALES)
    sheet = _sheet()
    assert sheet.cell(row=2, column=1).value == "North"
    assert sheet.cell(row=2, column=3).value == 120
    assert sheet.cell(row=2, column=4).value == 4.5, (
        "Only the new column should be written; the source data stays as it was."
    )


def test_saved_to_disk():
    """The change survives the program ending"""
    reset_sales()
    add_totals = require("add_totals")
    add_totals(SALES)
    fresh = load_workbook(SALES).active
    assert fresh.cell(row=2, column=5).value == 540.0, (
        "The workbook on disk has no totals in it. Assigning to a cell changes it in "
        "memory only — book.save(path) is what writes the file."
    )
