from openpyxl import load_workbook

SOURCE = "sales.xlsx"
DEST = "summary.xlsx"

EXPECTED = {"East": 300.0, "North": 967.5, "South": 1060.0, "West": 787.5}


def test_returns_revenue_by_region():
    """build_summary returns the revenue for each region"""
    reset_sales()
    build_summary = require("build_summary")
    result = build_summary(SOURCE, DEST)
    assert isinstance(result, dict), f"Expected a dict, got {type(result).__name__}."
    assert result == EXPECTED, f"Got {result}"


def test_total_row_excluded_from_return():
    """TOTAL belongs in the sheet, not in the returned dict"""
    reset_sales()
    build_summary = require("build_summary")
    assert "TOTAL" not in build_summary(SOURCE, DEST), (
        "The return value is regions only; the TOTAL row is a feature of the report."
    )


def test_workbook_written():
    """A new workbook appears, with the right sheet name"""
    reset_sales()
    require("build_summary")(SOURCE, DEST)
    book = load_workbook(DEST)
    assert book.sheetnames == ["Summary"], (
        f"Expected a single sheet named 'Summary', got {book.sheetnames}. "
        "Workbook() already gives you one sheet — rename it rather than adding another."
    )


def test_header_and_order():
    """Regions are listed alphabetically under a header"""
    reset_sales()
    require("build_summary")(SOURCE, DEST)
    sheet = load_workbook(DEST).active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("Region", "Units", "Revenue"), f"Row 1 is {rows[0]}"
    assert [row[0] for row in rows[1:5]] == ["East", "North", "South", "West"], (
        f"Got {[row[0] for row in rows[1:5]]}. Sort the regions so the report is stable."
    )


def test_region_numbers():
    """Units and revenue are summed per region"""
    reset_sales()
    require("build_summary")(SOURCE, DEST)
    rows = list(load_workbook(DEST).active.iter_rows(min_row=2, values_only=True))
    by_region = {row[0]: (row[1], row[2]) for row in rows}
    assert by_region["North"] == (215, 967.5), f"North came out as {by_region['North']}"
    assert by_region["West"] == (210, 787.5), (
        f"West came out as {by_region['West']} — its February row has an empty Units cell."
    )


def test_total_row():
    """The last row totals everything above it"""
    reset_sales()
    require("build_summary")(SOURCE, DEST)
    rows = list(load_workbook(DEST).active.iter_rows(values_only=True))
    assert rows[-1][0] == "TOTAL", f"The last row starts with {rows[-1][0]!r}"
    assert rows[-1][1] == 705, f"Total units came out as {rows[-1][1]}"
    assert abs(rows[-1][2] - 3115.0) < 0.001, f"Total revenue came out as {rows[-1][2]}"


def test_header_is_bold():
    """Row 1 is bold"""
    reset_sales()
    require("build_summary")(SOURCE, DEST)
    sheet = load_workbook(DEST).active
    assert sheet["A1"].font.bold, (
        "A1 is not bold. Font(bold=True) has to be assigned to each header cell."
    )


def test_source_untouched():
    """The source workbook is left exactly as it was"""
    reset_sales()
    require("build_summary")(SOURCE, DEST)
    sheet = load_workbook(SOURCE).active
    assert sheet.max_column == 4, (
        f"The source now has {sheet.max_column} columns. This step writes a new file; "
        "it does not modify the one it read."
    )
