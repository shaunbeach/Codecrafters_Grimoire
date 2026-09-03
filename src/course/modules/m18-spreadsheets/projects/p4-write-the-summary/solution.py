from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


def build_summary(source, destination):
    sheet = load_workbook(source).active

    totals = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        region, _month, units, price = row[0], row[1], row[2] or 0, row[3] or 0
        entry = totals.setdefault(region, [0, 0.0])
        entry[0] += units
        entry[1] += units * price

    report = Workbook()
    out = report.active
    out.title = "Summary"
    out.append(["Region", "Units", "Revenue"])

    for region in sorted(totals):
        units, revenue = totals[region]
        out.append([region, units, revenue])

    out.append([
        "TOTAL",
        sum(units for units, _ in totals.values()),
        sum(revenue for _, revenue in totals.values()),
    ])

    for cell in out[1]:
        cell.font = Font(bold=True)

    report.save(destination)
    return {region: revenue for region, (_units, revenue) in totals.items()}
