from openpyxl import load_workbook

SOURCE = "sales.xlsx"
DEST = "invoice.xlsx"


def _build(region="North"):
    require("build_invoice")(SOURCE, DEST, region)
    return load_workbook(DEST).active


def test_returns_the_total():
    """build_invoice returns the invoice total"""
    reset_sales()
    total = require("build_invoice")(SOURCE, DEST, "North")
    assert abs(total - 967.5) < 0.001, f"North is 540.0 + 427.5; you returned {total!r}"
    assert isinstance(total, float), f"Expected a float, got {type(total).__name__}."


def test_sheet_name_and_heading():
    """The sheet is named Invoice and headed INVOICE"""
    reset_sales()
    sheet = _build()
    assert load_workbook(DEST).sheetnames == ["Invoice"], (
        f"Got {load_workbook(DEST).sheetnames}"
    )
    assert sheet["A1"].value == "INVOICE", f"A1 holds {sheet['A1'].value!r}"


def test_reference_block():
    """Region and reference are where the brief says"""
    reset_sales()
    sheet = _build()
    assert (sheet["A2"].value, sheet["B2"].value) == ("Region", "North"), (
        f"Row 2 holds {sheet['A2'].value!r}, {sheet['B2'].value!r}"
    )
    assert sheet["B3"].value == "INV-NORTH", (
        f"B3 holds {sheet['B3'].value!r} — the region goes in capitals."
    )


def test_blank_row_then_header():
    """Row 4 is empty and the header is on row 5"""
    reset_sales()
    sheet = _build()
    assert sheet["A4"].value is None, (
        f"A4 holds {sheet['A4'].value!r}; row 4 is deliberately empty."
    )
    assert [c.value for c in sheet[5]][:4] == ["Month", "Units", "Unit Price", "Amount"], (
        f"Row 5 holds {[c.value for c in sheet[5]][:4]}"
    )


def test_line_items():
    """One row per month, with the amount worked out"""
    reset_sales()
    sheet = _build()
    assert [c.value for c in sheet[6]][:4] == ["January", 120, 4.5, 540.0], (
        f"Row 6 holds {[c.value for c in sheet[6]][:4]}"
    )
    assert [c.value for c in sheet[7]][:4] == ["February", 95, 4.5, 427.5], (
        f"Row 7 holds {[c.value for c in sheet[7]][:4]}"
    )


def test_total_row():
    """The last row totals column D"""
    reset_sales()
    sheet = _build()
    last = [c.value for c in sheet[sheet.max_row]]
    assert last[0] == "TOTAL", f"The last row starts with {last[0]!r}"
    assert abs((last[3] or 0) - 967.5) < 0.001, f"The total column holds {last[3]!r}"


def test_blank_units_count_as_zero():
    """West has an empty Units cell and still invoices"""
    reset_sales()
    total = require("build_invoice")(SOURCE, DEST, "West")
    assert abs(total - 787.5) < 0.001, (
        f"West is 210 x 3.75 plus an empty row; you returned {total!r}"
    )


def test_formatting():
    """The title and the header row are bold"""
    reset_sales()
    sheet = _build()
    assert sheet["A1"].font.bold and sheet["A1"].font.size == 14, (
        "A1 should be bold and size 14."
    )
    assert sheet["A5"].font.bold, "The header row should be bold."


def test_unknown_region_raises():
    """Asking for a region that does not exist is an error"""
    reset_sales()
    build_invoice = require("build_invoice")
    try:
        build_invoice(SOURCE, DEST, "Atlantis")
    except ValueError as exc:
        assert "Atlantis" in str(exc), (
            f"The message was {str(exc)!r}; naming the region makes the error useful."
        )
        return
    raise AssertionError(
        "Building an invoice for a region with no sales should raise ValueError."
    )
