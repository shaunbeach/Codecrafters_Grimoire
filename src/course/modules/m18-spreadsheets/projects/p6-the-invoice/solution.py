from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


def build_invoice(source, destination, region):
    sheet = load_workbook(source).active
    rows = [r for r in sheet.iter_rows(min_row=2, values_only=True) if r[0] == region]
    if not rows:
        raise ValueError(f"No sales for {region}")

    book = Workbook()
    out = book.active
    out.title = "Invoice"

    out.append(["INVOICE"])
    out.append(["Region", region])
    out.append(["Reference", f"INV-{region.upper()}"])
    out.append([])
    out.append(["Month", "Units", "Unit Price", "Amount"])

    total = 0.0
    for _region, month, units, price in rows:
        units = units or 0
        amount = units * price
        total += amount
        out.append([month, units, price, amount])

    out.append(["TOTAL", None, None, total])

    out["A1"].font = Font(bold=True, size=14)
    for cell in out[5]:
        cell.font = Font(bold=True)

    book.save(destination)
    return float(total)
