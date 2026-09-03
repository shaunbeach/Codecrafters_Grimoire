from openpyxl import load_workbook


def sheet_names(path):
    book = load_workbook(path)
    return book.sheetnames


def read_cell(path, sheet, ref):
    book = load_workbook(path)
    return book[sheet][ref].value
